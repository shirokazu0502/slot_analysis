import openpyxl
from openpyxl.styles import PatternFill

# カラムの制限をなくす（全てのカラムを表示）
pd.set_option('display.max_columns', None)

# 日付の末尾がイベ日である行を抽出
target_dates = main_data[main_data['day_end'] == 0]['日付']
my_df = main_data[main_data['機種名'] == "マイジャグラーV"]
filtered_my_df = my_data[my_df['日付'].isin(target_dates) | my_df['日付'].isin(target_dates - pd.DateOffset(
    days=1)) | my_df['日付'].isin(target_dates - pd.DateOffset(days=2)) | my_df['日付'].isin(target_dates - pd.DateOffset(days=3))]
filtered_my_df_event = my_df[my_df['日付'].isin(target_dates)]
filtered_my_df = filtered_my_df.copy()
filtered_my_df_event = filtered_my_df_event.copy()
filtered_my_df['日付'] = filtered_my_df['日付'].dt.strftime('%Y年%m月%d日(%a)')
filtered_my_df_event['日付'] = filtered_my_df_event['日付'].dt.strftime('%Y年%m月%d日(%a)')

pivot_table_all_samai = my_df.pivot_table(
    index = '日付', columns = '台番号', values = '差枚')

pivot_table_all_setting = my_df.pivot_table(
    index = '日付', columns = '台番号', values = 'setting')
pivot_table_sorted_all_setting = pivot_table_all_setting.sort_values(
    by = '日付', ascending = False)

pivot_table_sorted_all_samai = pivot_table_all_samai.sort_values(
    by = '日付', ascending = False)

pivot_table_all_with_totals = pivot_table_sorted_all_samai.copy()
pivot_table_all_with_totals.loc['合計'] = pivot_table_all_with_totals.sum()

pivot_table_samai = filtered_my_df.pivot_table(
    index = '日付', columns = '台番号', values = '差枚')

pivot_table_sorted_samai = pivot_table_samai.sort_values(
    by = '日付', ascending = False)

pivot_table_samai_event = filtered_my_df_event.pivot_table(
    index = '日付', columns = '台番号', values = '差枚')
pivot_table_sorted_samai_event = pivot_table_samai_event.sort_values(
    by = '日付', ascending = False)

pivot_table_event_with_totals = pivot_table_sorted_samai_event.copy()
pivot_table_event_with_totals.loc['合計'] = pivot_table_event_with_totals.sum()

pivot_table_game_count = filtered_my_df.pivot_table(
    index = '日付', columns = '台番号', values = 'G数')
pivot_table_sorted_game_count = pivot_table_game_count.sort_values(
    by = '日付', ascending = False)

pivot_table_gassan = filtered_my_df.pivot_table(
    index = '日付', columns = '台番号', values = '合成確率')
pivot_table_sorted_gassan = pivot_table_gassan.sort_values(
    by = '日付', ascending = False)

pivot_table_reg = filtered_my_df.pivot_table(
    index = '日付', columns = '台番号', values = 'RB確率')
pivot_table_sorted_reg = pivot_table_reg.sort_values(by='日付', ascending=False)

pivot_table_setting = filtered_my_df.pivot_table(
    index = '日付', columns = '台番号', values = 'setting')
pivot_table_sorted_setting = pivot_table_setting.sort_values(
    by = '日付', ascending = False)

pivot_table_setting_event = filtered_my_df_event.pivot_table(
    index = '日付', columns = '台番号', values = 'setting')
pivot_table_sorted_setting_event = pivot_table_setting_event.sort_values(
    by = '日付', ascending = False)
pivot_table_sorted_setting_event = pd.DataFrame(
    pivot_table_sorted_setting_event)
display(pivot_table_sorted_setting_event)
setting = 5
count_of_values_above_condition = pivot_table_sorted_setting_event.applymap(
    lambda x: 1 if x >= setting else 0).sum()
# print(type(count_of_values_above_condition))
# 新しい行を追加
pivot_table_sorted_setting_event = pd.concat(
    [pivot_table_sorted_setting_event, count_of_values_above_condition.to_frame().T], ignore_index = False)
pivot_table_sorted_setting_event.index = pivot_table_sorted_setting_event.index.to_list()[
    : -1] + ['高設定数']

# Excelファイルに保存
with pd.ExcelWriter(f'{shop_name}-data.xlsx', engine='openpyxl') as writer:
    pivot_table_all_samai.to_excel(
        writer, sheet_name = '全データの差枚', index = True)
    pivot_table_all_setting.to_excel(
        writer, sheet_name = '全データの設定', index = True)

    filtered_my_df.to_excel(writer, sheet_name='三日前まで含めた全項目データ', index=False)

    pivot_table_event_with_totals.to_excel(
        writer, sheet_name = 'eventのみの差枚', index = True)
    pivot_table_sorted_setting_event.to_excel(
        writer, sheet_name = 'eventのみの設定', index = True)

    pivot_table_sorted_samai.to_excel(
        writer, sheet_name = '三日前まで含めた差枚', index = True)

    pivot_table_sorted_game_count.to_excel(
        writer, sheet_name = '三日前まで含めたG数', index = True)

    pivot_table_sorted_reg.to_excel(
        writer, sheet_name = '三日前まで含めたRB確率', index = True)

    pivot_table_sorted_gassan.to_excel(
        writer, sheet_name = '三日前まで含めた合算確率', index = True)

    pivot_table_sorted_setting.to_excel(
        writer, sheet_name = '三日前まで含めたsetting', index = True)

    workbook = writer.book
    worksheet_all_samai = writer.sheets['全データの差枚']
    worksheet_all_setting = writer.sheets['全データの設定']
    worksheet_setting = writer.sheets['三日前まで含めたsetting']
    worksheet_samai = writer.sheets['三日前まで含めた差枚']
    worksheet_game_count = writer.sheets['三日前まで含めたG数']
    worksheet_reg_probability = writer.sheets['三日前まで含めたRB確率']
    worksheet_total_probability = writer.sheets['三日前まで含めた合算確率']
    worksheet_setting_event = writer.sheets['eventのみの設定']
    worksheet_samai_event = writer.sheets['eventのみの差枚']

    # 値が5以上のセルを赤く塗りつぶす
    red_fill = PatternFill(start_color='FFFF0000',
                           end_color = 'FFFF0000', fill_type = 'solid')
    # 全データ
    for cell in worksheet_all_setting.iter_rows(min_row=2, min_col=2, max_row=len(pivot_table_sorted_all_setting) + 1, max_col=len(pivot_table_sorted_all_setting.columns) + 1):
        for cell in cell:
            if isinstance(cell.value, (int, float)) and cell.value >= 5:
                cell.fill = red_fill
                high_setting_row, high_setting_column = cell.row, cell.column
                worksheet_all_samai.cell(high_setting_row, high_setting_column).fill = red_fill

    # イベントの三日前まで含めた
    for cell in worksheet_setting.iter_rows(min_row=2, min_col=2, max_row=len(pivot_table_setting) + 1, max_col=len(pivot_table_sorted_setting.columns) + 1):
        for cell in cell:
            if isinstance(cell.value, (int, float)) and cell.value >= 5:
                cell.fill = red_fill
                high_setting_row, high_setting_column = cell.row, cell.column
                worksheet_samai.cell(
                    high_setting_row, high_setting_column).fill = red_fill
                worksheet_game_count.cell(
                    high_setting_row, high_setting_column).fill = red_fill
                worksheet_reg_probability.cell(
                    high_setting_row, high_setting_column).fill = red_fill
                worksheet_total_probability.cell(
                    high_setting_row, high_setting_column).fill = red_fill

    # イベントのみ
    for cell in worksheet_setting_event.iter_rows(min_row=2, min_col=2, max_row=len(pivot_table_setting) + 1, max_col=len(pivot_table_sorted_setting_event.columns) + 1):
        for cell in cell:
            if isinstance(cell.value, (int, float)) and cell.value >= 5:
                cell.fill = red_fill
                high_setting_row, high_setting_column = cell.row, cell.column
                worksheet_samai_event.cell(
                    high_setting_row, high_setting_column).fill = red_fill


# pivot_table_sorted_setting.to_csv(f"{shop_name}-setting.csv", encoding="shift-jis")
