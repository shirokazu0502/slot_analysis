import pandas as pd
from matplotlib import pyplot as plt
import japanize_matplotlib
import openpyxl
from openpyxl.styles import PatternFill
pd.set_option('display.max_rows', 10000)
shop_name="ミリオン北島店"
main_data=pd.read_csv(f"{shop_name}.csv", encoding="shift-jis")

#分数を変換
main_data['合成確率'] = main_data['合成確率'] .str.replace('1/','')
main_data['合成確率']=main_data['合成確率'].astype('float64')
main_data['BB確率'] = main_data['BB確率'] .str.replace('1/','')
main_data['BB確率']=main_data['BB確率'].astype('float64')
main_data['RB確率'] = main_data['RB確率'] .str.replace('1/','')
main_data['RB確率']=main_data['RB確率'].astype('float64')

#一旦すべて設定１とする
main_data['setting'] = 1
#アイムジャグラー推定設定
main_data.loc[(main_data['RB確率'] <= 399.6) & (main_data['機種名']=="アイムジャグラーEX-TP"),'setting'] = 2
main_data.loc[(main_data['RB確率'] <= 331.0) & (main_data['機種名']=="アイムジャグラーEX-TP"),'setting'] = 3
main_data.loc[(main_data['RB確率'] <= 315.1) & (main_data['機種名']=="アイムジャグラーEX-TP"),'setting'] = 4
main_data.loc[(main_data['RB確率'] <= 255.0) & (main_data['機種名']=="アイムジャグラーEX-TP"),'setting'] = 5
main_data.loc[(main_data['RB確率'] <= 255.0) & (main_data['機種名']=="アイムジャグラーEX-TP") & (main_data["ブドウ確率"]<5.90),'setting'] = 6
main_data.loc[(main_data['G数'] < 4000),'setting'] = 0 #回転数が4000未満は設定0として扱う
#マイジャグラー推定設定
main_data.loc[(main_data['RB確率'] <= 385.5) & (main_data['機種名']=="マイジャグラーV") & (main_data["ブドウ確率"]<=5.90),'setting'] = 2
main_data.loc[(main_data['RB確率'] <= 336.1) & (main_data['機種名']=="マイジャグラーV") & (main_data["ブドウ確率"]<=5.90),'setting'] = 3
main_data.loc[(main_data['RB確率'] <= 290.0) & (main_data['機種名']=="マイジャグラーV") & (main_data["ブドウ確率"]<=5.90),'setting'] = 4
main_data.loc[(main_data['RB確率'] <= 268.6) & (main_data['機種名']=="マイジャグラーV") & (main_data["ブドウ確率"]<=5.85),'setting'] = 5
main_data.loc[(main_data['RB確率'] <= 229.1) & (main_data['機種名']=="マイジャグラーV") & (main_data["ブドウ確率"]<=5.80),'setting'] = 6
main_data.loc[(main_data['G数'] < 4000),'setting'] = 0 #回転数が4000未満は設定0として扱う
#ファンキージャグラー推定設定
main_data.loc[(main_data['RB確率'] <= 407.1) & (main_data['機種名']=="ファンキージャグラー2") & (main_data["ブドウ確率"]<=5.92),'setting'] = 2
main_data.loc[(main_data['RB確率'] <= 366.1) & (main_data['機種名']=="ファンキージャグラー2") & (main_data["ブドウ確率"]<=5.92),'setting'] = 3
main_data.loc[(main_data['RB確率'] <= 322.8) & (main_data['機種名']=="ファンキージャグラー2") & (main_data["ブドウ確率"]<=5.92),'setting'] = 4
main_data.loc[(main_data['RB確率'] <= 299.3) & (main_data['機種名']=="ファンキージャグラー2") & (main_data["ブドウ確率"]<=5.88),'setting'] = 5
main_data.loc[(main_data['RB確率'] <= 262.1) & (main_data['機種名']=="ファンキージャグラー2") & (main_data["ブドウ確率"]<=5.83),'setting'] = 6
main_data.loc[(main_data['G数'] < 4000),'setting'] = 0 #回転数が4000未満は設定0として扱う
#ハッピージャグラー推定設定
main_data.loc[(main_data['RB確率'] <= 362.08) & (main_data['機種名']=="ハッピージャグラーVIII") & (main_data["ブドウ確率"]<=6.04),'setting'] = 2
main_data.loc[(main_data['RB確率'] <= 332.67) & (main_data['機種名']=="ハッピージャグラーVIII") & (main_data["ブドウ確率"]<=6.04),'setting'] = 3
main_data.loc[(main_data['RB確率'] <= 300.62) & (main_data['機種名']=="ハッピージャグラーVIII") & (main_data["ブドウ確率"]<=6.04),'setting'] = 4
main_data.loc[(main_data['RB確率'] <= 273.07) & (main_data['機種名']=="ハッピージャグラーVIII") & (main_data["ブドウ確率"]<=6.01),'setting'] = 5
main_data.loc[(main_data['RB確率'] <= 256.00) & (main_data['機種名']=="ハッピージャグラーVIII") & (main_data["ブドウ確率"]<=5.98),'setting'] = 6
main_data.loc[(main_data['G数'] < 4000),'setting'] = 0 #回転数が4000未満は設定0として扱う

#時刻データに変換
main_data['日付'] = pd.to_datetime(main_data['日付'])
main_data['day_end'] = main_data['日付'].astype(str).str[-1].astype(int) #日付末尾


# 前日の差枚、G数、BB確率、RB確率を計算して新しい列に追加
main_data['前日の差枚'] = main_data.groupby("台番号")['差枚'].shift(-1)
main_data['前日のG数'] = main_data.groupby("台番号")['G数'].shift(-1)
main_data['前日のBB確率'] = main_data.groupby("台番号")['BB確率'].shift(-1)
main_data['前日のRB確率'] = main_data.groupby("台番号")['RB確率'].shift(-1)
main_data.to_csv(f"{shop_name}-update.csv", index=False, encoding = "shift-jis")

#設定の比率
n_target0, n_target1,n_target2,n_target3,n_target4,n_target5,n_target6= len(main_data[main_data['setting'] == 0]), len(main_data[main_data['setting'] == 1]) , len(main_data[main_data['setting'] == 2]), len(main_data[main_data['setting'] == 3]), len(main_data[main_data['setting'] == 4]), len(main_data[main_data['setting'] == 5]), len(main_data[main_data['setting'] == 6])
n_all = n_target0+n_target1+n_target2+n_target3+n_target4+n_target5+n_target6
print('回転数不足 の割合 :', n_target0/n_all) # target0の割合
print('設定1 の割合 :', n_target1/n_all) # target1の割合
print('設定2 の割合 :', n_target2/n_all) # target2の割合
print('設定3 の割合 :', n_target3/n_all) # target3の割合
print('設定4 の割合 :', n_target4/n_all) # target4の割合
print('設定5 の割合 :', n_target5/n_all) # target5の割合
print('設定6 の割合 :', n_target6/n_all) # target6の割合

# サブプロットを配置
fig, axes = plt.subplots(3, 2, figsize=(8, 6))

# ヒストグラムを作成し配置
axes[0, 0].hist(main_data["setting"], bins=20, color='blue', alpha=0.7)
axes[0, 0].set_xlabel("setting")
axes[0, 0].set_ylabel("number of setting")
axes[0, 0].set_title("設定投入数")

#全データの中で設定5以上の投入率を調べる
# "setting" 列が5以上の行を抽出
setting_upperfive_df = main_data[main_data['setting'] >= 5]
# "day_end" 列ごとに条件を満たす行の数を計算
upperfize_by_day = setting_upperfive_df.groupby('day_end')['setting'].count()
all_by_day = main_data.groupby('day_end')['setting'].count()

percentage_by_day = upperfize_by_day*100/all_by_day
# バーの位置を設定
x = range(len(percentage_by_day))

# バーの高さを設定
y = percentage_by_day.values
axes[0, 1].bar(x, y)
axes[0, 1].set_xlabel('day_end')
axes[0, 1].set_ylabel("%")
axes[0, 1].set_title('percentage of setting >= 5 by day_end')
plt.xticks(rotation=45)  # X軸のラベルを回転させて可読性を向上させる

#アイムジャグラーの高設定投入率を調べる
im_all_df = main_data[main_data["機種名"]=="アイムジャグラーEX-TP"]
im_upperfive_df = im_all_df[im_all_df["setting"]>=5]

upperfize_by_day = im_upperfive_df.groupby('day_end')['setting'].count()
all_by_day = im_all_df.groupby('day_end')['setting'].count()

percentage_by_day =upperfize_by_day*100/all_by_day
# バーの位置を設定
x = range(len(percentage_by_day))

# バーの高さを設定
y = percentage_by_day.values
axes[1, 0].bar(x, y)
axes[1, 0].set_xlabel('day_end')
axes[1, 0].set_ylabel("%")
axes[1, 0].set_title('アイムジャグラー設定5以上投入率')

#マイジャグラーの高設定投入率を調べる
my_all_df = main_data[main_data["機種名"]=="マイジャグラーV"]
my_upperfive_df = my_all_df[my_all_df["setting"]>=5]

upperfize_by_day = my_upperfive_df.groupby('day_end')['setting'].count()
all_by_day = my_all_df.groupby('day_end')['setting'].count()

percentage_by_day =upperfize_by_day*100/all_by_day
# バーの位置を設定
x = range(len(percentage_by_day))

# バーの高さを設定
y = percentage_by_day.values
axes[1, 1].bar(x, y)
axes[1, 1].set_xlabel('day_end')
axes[1, 1].set_ylabel("%")
axes[1, 1].set_title('マイジャグラー設定5以上投入率')

#ファンキージャグラーの高設定投入率を調べる
fanky_all_df = main_data[main_data["機種名"]=="ファンキージャグラー2"]
fanky_upperfive_df = fanky_all_df[fanky_all_df["setting"]>=5]

upperfize_by_day = fanky_upperfive_df.groupby('day_end')['setting'].count()
all_by_day = fanky_all_df.groupby('day_end')['setting'].count()

percentage_by_day =upperfize_by_day*100/all_by_day

# バーの位置を設定
x = range(len(percentage_by_day))

# バーの高さを設定
y = percentage_by_day.values
axes[2, 0].bar(x, y)
axes[2, 0].set_xlabel('day_end')
axes[2, 0].set_ylabel("%")
axes[2, 0].set_title('ファンキージャグラー設定5以上投入率')

#ハッピージャグラーの高設定投入率を調べる
happy_all_df = main_data[main_data["機種名"]=="ハッピージャグラーVIII"]
happy_upperfive_df = happy_all_df[happy_all_df["setting"]>=5]

upperfize_by_day = happy_upperfive_df.groupby('day_end')['setting'].count()
all_by_day = happy_all_df.groupby('day_end')['setting'].count()

percentage_by_day =upperfize_by_day*100/all_by_day
# バーの位置を設定
x = range(len(percentage_by_day))

# バーの高さを設定
y = percentage_by_day.values
axes[2, 1].bar(x, y)
axes[2, 1].set_xlabel('day_end')
axes[2, 1].set_ylabel("%")
axes[2, 1].set_title('ハッピージャグラー設定5以上投入率')
# サブプロット間の横方向のスペースを調整
plt.subplots_adjust(wspace=0.3)  # 各サブプロット間のスペースを調整
# サブプロット間のスペースを調整
plt.tight_layout()

# 図を保存
plt.savefig(f'{shop_name}.png', dpi=300)
plt.show()

# カラムの制限をなくす（全てのカラムを表示）
pd.set_option('display.max_columns', None)

# 日付の末尾が8である行を抽出
target_dates = main_data[main_data['day_end'] == 3]['日付']
filtered_df = main_data[main_data['日付'].isin(target_dates) | main_data['日付'].isin(target_dates - pd.DateOffset(days=1))| main_data['日付'].isin(target_dates - pd.DateOffset(days=2))]
filtered_df_event = main_data[main_data['日付'].isin(target_dates)]
filtered_df = filtered_df.copy()
filtered_df_event = filtered_df_event.copy()
filtered_df['日付'] = filtered_df['日付'].dt.strftime('%Y年%m月%d日(%a)')
filtered_df_event['日付'] = filtered_df_event['日付'].dt.strftime('%Y年%m月%d日(%a)')
filtered_my_df = filtered_df[filtered_df['機種名'] == "アイムジャグラーEX-TP"]
filtered_my_df_event = filtered_df_event[filtered_df_event['機種名'] == "アイムジャグラーEX-TP"]

# # 台番号カラムから一意なレコードを抽出してリストに格納
# machine_nums = main_data['台番号'].drop_duplicates().tolist()
pivot_table_samai = filtered_my_df.pivot_table(index='日付', columns='台番号', values='差枚')
pivot_table_sorted_samai = pivot_table_samai.sort_values(by='日付', ascending=False)

pivot_table_samai_event = filtered_my_df_event.pivot_table(index='日付', columns='台番号', values='差枚')
pivot_table_sorted_samai_event = pivot_table_samai_event.sort_values(by='日付', ascending=False)

pivot_table_event_with_totals = pivot_table_sorted_samai_event.copy()
pivot_table_event_with_totals.loc['合計'] = pivot_table_event_with_totals.sum()

pivot_table_gassan = filtered_my_df.pivot_table(index='日付', columns='台番号', values='合成確率')
pivot_table_sorted_gassan = pivot_table_gassan.sort_values(by='日付', ascending=False)

pivot_table_setting=filtered_my_df.pivot_table(index='日付', columns='台番号', values='setting')
pivot_table_sorted_setting = pivot_table_setting.sort_values(by='日付', ascending=False)

pivot_table_setting_event=filtered_my_df_event.pivot_table(index='日付', columns='台番号', values='setting')
pivot_table_sorted_setting_event=pivot_table_setting_event.sort_values(by='日付', ascending=False)
pivot_table_sorted_setting_event=pd.DataFrame(pivot_table_sorted_setting_event)
setting = 5
count_of_values_above_condition = pivot_table_sorted_setting_event.applymap(lambda x: 1 if x >= setting else 0).sum()
# print(type(count_of_values_above_condition))
# 新しい行を追加
pivot_table_sorted_setting_event = pd.concat([pivot_table_sorted_setting_event, count_of_values_above_condition.to_frame().T], ignore_index=False)
pivot_table_sorted_setting_event.index = pivot_table_sorted_setting_event.index.to_list()[:-1] + ['高設定数']

# Excelファイルに保存
with pd.ExcelWriter(f'{shop_name}-data.xlsx', engine='openpyxl') as writer:
    filtered_my_df.to_excel(writer, sheet_name='二日前まで含めた全項目データ', index=False)
    pivot_table_event_with_totals.to_excel(writer, sheet_name='eventのみの差枚', index=True)
    pivot_table_sorted_setting_event.to_excel(writer, sheet_name='eventのみの設定', index=True)
    pivot_table_sorted_samai.to_excel(writer, sheet_name='二日前まで含めた差枚', index=True)
    pivot_table_sorted_gassan.to_excel(writer, sheet_name='二日前まで含めた合算確率', index=True)
    pivot_table_sorted_setting.to_excel(writer, sheet_name='二日前まで含めたsetting', index=True)
    workbook = writer.book
    worksheet_setting = writer.sheets['二日前まで含めたsetting']
    worksheet_samai = writer.sheets['二日前まで含めた差枚']
    worksheet_total_probability=writer.sheets['二日前まで含めた合算確率']
    
    # 値が5以上のセルを赤く塗りつぶす
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    for cell in worksheet_setting.iter_rows(min_row=2, min_col=2, max_row=len(pivot_table_setting) + 1, max_col=len(pivot_table_sorted_setting.columns) + 1):
        for cell in cell:
            if isinstance(cell.value, (int, float)) and cell.value >= 5:
                cell.fill = red_fill
                high_setting_row, high_setting_column = cell.row, cell.column
                worksheet_samai.cell(high_setting_row, high_setting_column).fill=red_fill
                worksheet_total_probability.cell(high_setting_row, high_setting_column).fill=red_fill





# pivot_table_sorted_setting.to_csv(f"{shop_name}-setting.csv", encoding="shift-jis")